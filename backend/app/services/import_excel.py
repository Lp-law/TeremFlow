from __future__ import annotations

import datetime as dt
from decimal import Decimal
from typing import Any

from fastapi import HTTPException
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.enums import CaseType, FeeEventType
from app.models.case import Case
from app.services.cases import create_case, update_case_from_excel, merge_raw_import_fields, EXCEL_UPDATE_FIELDS

# Logical fields that map to Case and drive billing/calculations. All other columns go to raw_import_fields_json.
OPERATIONAL_FIELDS = EXCEL_UPDATE_FIELDS | {"case_reference"}


def _norm(s: Any) -> str:
    return str(s).strip().replace("\u200f", "").replace("\u200e", "").lower()


KNOWN_COLUMNS: dict[str, str] = {
    # case reference
    "case": "case_reference",
    "case_reference": "case_reference",
    "תיק": "case_reference",
    "מספר תיק": "case_reference",
    # case name (plaintiff / display name)
    "case_name": "case_name",
    "שם תיק": "case_name",
    "שם התיק": "case_name",
    # type
    "case_type": "case_type",
    "סוג תיק": "case_type",
    "סוג": "case_type",
    # open date
    "open_date": "open_date",
    "תאריך פתיחה": "open_date",
    "פתיחה": "open_date",
    # deductible
    "deductible_usd": "deductible_usd",
    "אקסס usd": "deductible_usd",
    "excess usd": "deductible_usd",
    "deductible_ils": "deductible_ils_gross",
    "deductible_ils_gross": "deductible_ils_gross",
    "אקסס שח": "deductible_ils_gross",
    "אקסס ש\"ח": "deductible_ils_gross",
    # branch (Excel column B)
    "branch": "branch_name",
    "branch_name": "branch_name",
    "סניף": "branch_name",
    # retainer anchor (Excel column C)
    "retainer_anchor": "retainer_anchor_date",
    "retainer_anchor_date": "retainer_anchor_date",
    "תאריך עוגן": "retainer_anchor_date",
    # Excel H: total retainer paid to date (snapshot)
    "h": "retainer_snapshot_ils_gross",
    "retainer_snapshot": "retainer_snapshot_ils_gross",
    "retainer_snapshot_ils_gross": "retainer_snapshot_ils_gross",
    "ריטיינר שולם": "retainer_snapshot_ils_gross",
    "total retainer": "retainer_snapshot_ils_gross",
    'סה"כ שולם בריטיינר': "retainer_snapshot_ils_gross",
    # retainer snapshot through month (YYYY-MM-01) — last month included in H. If omitted and H set, default = last month.
    "retainer_snapshot_through_month": "retainer_snapshot_through_month",
    "חודש סיום ריטיינר": "retainer_snapshot_through_month",
    "snapshot_through": "retainer_snapshot_through_month",
    # Excel I: total non-attorney expenses (snapshot)
    "i": "expenses_snapshot_ils_gross",
    "expenses_snapshot": "expenses_snapshot_ils_gross",
    "expenses_snapshot_ils_gross": "expenses_snapshot_ils_gross",
    "הוצאות אחרות": "expenses_snapshot_ils_gross",
    "other expenses": "expenses_snapshot_ils_gross",
    'סה"כ הוצאות בתיק': "expenses_snapshot_ils_gross",
    # Historical fee stages: comma-separated FeeEventType codes
    "historical_fee_stages": "historical_fee_stages",
    "שלבי שכ״ט עבר": "historical_fee_stages",
    # Legacy free-text (e.g. "פירוט חיוב שכ״ט עו״ד") — read-only, not parsed
    "legacy_fee_text": "legacy_fee_text",
    "פירוט חיוב שכ״ט עו״ד": "legacy_fee_text",
    'פירוט חיוב שכ"ט עו"ד': "legacy_fee_text",
    "fee_charges_raw": "legacy_fee_text",
    # Optional: prefill performed_fee_stage_codes (comma-separated); unknown codes ignored with warning
    "performed_fee_stage_codes": "performed_fee_stage_codes",
    "שלבים שבוצעו": "performed_fee_stage_codes",
}

VALID_FEE_EVENT_TYPES = frozenset(e.value for e in FeeEventType)
# STAGE_BILLING is composite only; do not allow in historical_fee_stages
ALLOWED_HISTORICAL_FEE_STAGES = frozenset(e.value for e in FeeEventType if e != FeeEventType.STAGE_BILLING)
# Same set for performed_fee_stage_codes (stage-billing codes that have rates)
ALLOWED_PERFORMED_CODES = ALLOWED_HISTORICAL_FEE_STAGES


def _parse_date(v: Any) -> dt.date:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    # try ISO
    try:
        return dt.date.fromisoformat(str(v))
    except Exception:
        raise ValueError(f"Invalid date: {v}")


def _parse_decimal_ge_zero(v: Any, field_name: str) -> Decimal | None:
    """Parse decimal >= 0. Returns None for empty/None."""
    if v is None or v == "":
        return None
    try:
        d = Decimal(str(v))
        if d < 0:
            raise ValueError(f"{field_name} must be >= 0")
        return d
    except Exception as e:
        raise ValueError(f"Invalid {field_name}: {v}") from e


def _parse_historical_fee_stages(v: Any) -> list[str] | None:
    """Parse comma-separated FeeEventType codes. Empty/blank -> None. Invalid code -> raise."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    s = str(v).strip()
    if not s:
        return None
    parts = [p.strip() for p in s.split(",") if p.strip()]
    if not parts:
        return None
    invalid = [p for p in parts if p not in ALLOWED_HISTORICAL_FEE_STAGES]
    if invalid:
        raise ValueError(f"historical_fee_stages: קוד לא מוכר: {invalid[0]}. קודים חוקיים: {', '.join(sorted(ALLOWED_HISTORICAL_FEE_STAGES))}")
    return parts


def _parse_performed_fee_stage_codes(v: Any) -> tuple[list[str] | None, list[str]]:
    """
    Parse comma-separated performed_fee_stage_codes. Returns (valid_codes, unknown_codes).
    Unknown codes are ignored (not added to valid); valid can be empty. Does not raise.
    """
    if v is None or (isinstance(v, str) and not v.strip()):
        return None, []
    s = str(v).strip()
    if not s:
        return None, []
    parts = [p.strip() for p in s.split(",") if p.strip()]
    if not parts:
        return None, []
    valid = [p for p in parts if p in ALLOWED_PERFORMED_CODES]
    unknown = [p for p in parts if p not in ALLOWED_PERFORMED_CODES]
    return (valid if valid else None), unknown


def _parse_raw_value(v: Any) -> Any:
    """Parse cell to a scalar for raw_import_fields_json: number, date, bool, or string. Empty -> None. No raise."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.isoformat() if hasattr(v, "isoformat") else str(v)
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return v
    s = str(v).strip()
    if s == "":
        return None
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    try:
        return dt.date.fromisoformat(s)
    except Exception:
        pass
    if s.lower() in ("true", "yes", "1"):
        return True
    if s.lower() in ("false", "no", "0"):
        return False
    return s


def _parse_case_type(v: Any) -> CaseType:
    s = _norm(v)
    mapping = {
        "court": CaseType.COURT,
        "תיק ביהמ\"ש": CaseType.COURT,
        "תיק בית המשפט": CaseType.COURT,
        "ביהמ\"ש": CaseType.COURT,
        "demand_letter": CaseType.DEMAND_LETTER,
        "מכתב דרישה": CaseType.DEMAND_LETTER,
        "small_claims": CaseType.SMALL_CLAIMS,
        "תביעות קטנות": CaseType.SMALL_CLAIMS,
    }
    if s in mapping:
        return mapping[s]
    # try enum literal
    try:
        return CaseType(str(v))
    except Exception:
        raise ValueError(f"Invalid case_type: {v}")


def _parse_data_to_updates(data: dict) -> dict:
    """
    Parse row data into a dict of logical field -> value for update-import.
    None = empty/invalid (caller uses overwrite_blanks to decide whether to clear).
    Raises ValueError for missing case_reference or invalid provided values.
    """
    out: dict[str, Any] = {}
    ref = str(data.get("case_reference") or "").strip()
    if not ref:
        raise ValueError("Missing case_reference")
    out["case_reference"] = ref

    out["case_name"] = (str(data.get("case_name") or "").strip() or None)
    if data.get("case_type") not in (None, ""):
        out["case_type"] = _parse_case_type(data["case_type"])
    else:
        out["case_type"] = None
    if data.get("open_date") not in (None, ""):
        out["open_date"] = _parse_date(data["open_date"])
    else:
        out["open_date"] = None
    out["branch_name"] = (str(data.get("branch_name") or "").strip() or None)
    if data.get("deductible_usd") not in (None, ""):
        out["deductible_usd"] = _parse_decimal_ge_zero(data["deductible_usd"], "deductible_usd")
    else:
        out["deductible_usd"] = None
    if data.get("deductible_ils_gross") not in (None, ""):
        out["deductible_ils_gross"] = _parse_decimal_ge_zero(data["deductible_ils_gross"], "deductible_ils_gross")
    else:
        out["deductible_ils_gross"] = None
    if data.get("retainer_anchor_date") not in (None, ""):
        out["retainer_anchor_date"] = _parse_date(data["retainer_anchor_date"])
    else:
        out["retainer_anchor_date"] = None
    out["retainer_snapshot_ils_gross"] = _parse_decimal_ge_zero(data.get("retainer_snapshot_ils_gross"), "retainer_snapshot_ils_gross")
    if data.get("retainer_snapshot_through_month") not in (None, ""):
        out["retainer_snapshot_through_month"] = _parse_date(data["retainer_snapshot_through_month"])
    else:
        out["retainer_snapshot_through_month"] = None
    out["expenses_snapshot_ils_gross"] = _parse_decimal_ge_zero(data.get("expenses_snapshot_ils_gross"), "expenses_snapshot_ils_gross")
    out["historical_fee_stages"] = _parse_historical_fee_stages(data.get("historical_fee_stages"))
    raw = data.get("legacy_fee_text")
    out["legacy_fee_text"] = (str(raw).strip() or None) if raw not in (None, "") else None
    performed_valid, _ = _parse_performed_fee_stage_codes(data.get("performed_fee_stage_codes"))
    out["performed_fee_stage_codes"] = performed_valid
    return out


def _build_col_map_and_rows(file_bytes: bytes) -> tuple[list[tuple], dict[int, str], list[tuple]]:
    """Load workbook, return (header_row, col_map, data_rows). Raises HTTPException if empty or missing required for create."""
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty Excel file")
    header = rows[0]
    col_map: dict[int, str] = {}
    for idx, name in enumerate(header):
        key = _norm(name)
        if key in KNOWN_COLUMNS:
            col_map[idx] = KNOWN_COLUMNS[key]
        else:
            key2 = key.replace(" ", "")
            if key2 in KNOWN_COLUMNS:
                col_map[idx] = KNOWN_COLUMNS[key2]
    return header, col_map, rows[1:]


def import_cases_from_excel(db: Session, file_bytes: bytes) -> dict:
    _, col_map, data_rows = _build_col_map_and_rows(file_bytes)
    required = {"case_reference", "case_type", "open_date"}
    if not required.issubset(set(col_map.values())):
        raise HTTPException(status_code=400, detail=f"Missing required columns. Need at least: {sorted(required)}")

    created = 0
    skipped_empty_rows = 0
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    for r_i, row in enumerate(data_rows, start=2):
        if not any(row):
            skipped_empty_rows += 1
            continue
        data: dict[str, Any] = {}
        for idx, field in col_map.items():
            data[field] = row[idx] if idx < len(row) else None
        try:
            payload = type("Obj", (), {})()
            payload.case_reference = str(data["case_reference"] or "").strip()
            if not payload.case_reference:
                raise ValueError("Missing case_reference")
            payload.case_name = (str(data.get("case_name") or "").strip() or None)
            payload.case_type = _parse_case_type(data["case_type"])
            payload.open_date = _parse_date(data["open_date"])
            payload.deductible_usd = Decimal(str(data["deductible_usd"])) if data.get("deductible_usd") not in (None, "") else None
            payload.deductible_ils_gross = (
                Decimal(str(data["deductible_ils_gross"])) if data.get("deductible_ils_gross") not in (None, "") else None
            )
            payload.branch_name = (str(data.get("branch_name") or "").strip() or None)
            payload.retainer_anchor_date = _parse_date(data["retainer_anchor_date"]) if data.get("retainer_anchor_date") not in (None, "") else None
            # Excel H, I: snapshots (>= 0)
            payload.retainer_snapshot_ils_gross = _parse_decimal_ge_zero(data.get("retainer_snapshot_ils_gross"), "retainer_snapshot_ils_gross")
            payload.expenses_snapshot_ils_gross = _parse_decimal_ge_zero(data.get("expenses_snapshot_ils_gross"), "expenses_snapshot_ils_gross")
            # snapshot_through_month: if H set and not in Excel, default = last month (accruals from this month)
            if payload.retainer_snapshot_ils_gross is not None:
                if data.get("retainer_snapshot_through_month") not in (None, ""):
                    payload.retainer_snapshot_through_month = _parse_date(data["retainer_snapshot_through_month"])
                else:
                    today = dt.date.today()
                    first_this_month = dt.date(today.year, today.month, 1)
                    last_day_prev = first_this_month - dt.timedelta(days=1)
                    payload.retainer_snapshot_through_month = dt.date(last_day_prev.year, last_day_prev.month, 1)
            payload.historical_fee_stages = _parse_historical_fee_stages(data.get("historical_fee_stages"))
            raw = data.get("legacy_fee_text")
            payload.legacy_fee_text = (str(raw).strip() or None) if raw not in (None, "") else None
            performed_valid, performed_unknown = _parse_performed_fee_stage_codes(data.get("performed_fee_stage_codes"))
            if performed_unknown:
                warnings.append({"row": r_i, "message": f"unknown performed_fee_stage_codes ignored: {performed_unknown}", "data": data})
            payload.performed_fee_stage_codes = performed_valid
            create_case(db, payload)
            created += 1
        except HTTPException as e:
            # Preserve meaningful API details (e.g. duplicates, BOI FX failures).
            errors.append({"row": r_i, "error": str(e.detail), "data": data})
        except Exception as e:
            errors.append({"row": r_i, "error": str(e), "data": data})

    return {
        "created": created,
        "skipped_empty_rows": skipped_empty_rows,
        "errors": errors[:50],
        "error_count": len(errors),
        "warnings": warnings[:50],
        "warning_count": len(warnings),
    }


def _build_raw_from_row(header: list, col_map: dict[int, str], row: tuple, operational: set[str]) -> dict[str, Any]:
    """Build raw_import_fields_json dict: keys = header names, values = parsed scalars. Only columns NOT in operational."""
    raw: dict[str, Any] = {}
    for idx in range(max(len(header), len(row))):
        header_name = str(header[idx]).strip() if idx < len(header) and header[idx] is not None else f"col_{idx}"
        value = row[idx] if idx < len(row) else None
        if idx in col_map and col_map[idx] in operational:
            continue
        raw[header_name] = _parse_raw_value(value)
    return raw


def import_cases_from_excel_update(db: Session, file_bytes: bytes, *, overwrite_blanks: bool = False) -> dict:
    """
    Update existing cases by case_reference. Same header/column mapping as create-import.
    Operational fields -> Case columns; all other columns -> raw_import_fields_json (display-only).
    Blank does not overwrite unless overwrite_blanks=True. No auto-create: missing case -> row error.
    """
    header, col_map, data_rows = _build_col_map_and_rows(file_bytes)
    if "case_reference" not in col_map.values():
        raise HTTPException(status_code=400, detail="Missing required column: case_reference (for lookup)")

    updated = 0
    skipped_empty_rows = 0
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []

    for r_i, row in enumerate(data_rows, start=2):
        if not any(row):
            skipped_empty_rows += 1
            continue
        data: dict[str, Any] = {}
        for idx, field in col_map.items():
            data[field] = row[idx] if idx < len(row) else None
        try:
            parsed = _parse_data_to_updates(data)
            ref = parsed["case_reference"]
            case = db.query(Case).filter(Case.case_reference == ref).first()
            if not case:
                errors.append({"row": r_i, "error": "Case not found for update", "data": data})
                continue
            updates = {}
            for k in EXCEL_UPDATE_FIELDS:
                if k not in parsed:
                    continue
                if parsed[k] is not None:
                    updates[k] = parsed[k]
                elif overwrite_blanks and k in data:
                    updates[k] = None
            raw = _build_raw_from_row(header, col_map, row, OPERATIONAL_FIELDS)
            if updates:
                update_case_from_excel(db, case, updates, overwrite_blanks=overwrite_blanks)
            if raw:
                merge_raw_import_fields(case, raw, overwrite_blanks=overwrite_blanks)
                db.commit()
                db.refresh(case)
            if updates or raw:
                updated += 1
        except HTTPException as e:
            errors.append({"row": r_i, "error": str(e.detail), "data": data})
        except Exception as e:
            errors.append({"row": r_i, "error": str(e), "data": data})

    return {
        "updated": updated,
        "created": 0,
        "skipped_empty_rows": skipped_empty_rows,
        "errors": errors[:50],
        "error_count": len(errors),
        "warnings": warnings[:50],
        "warning_count": len(warnings),
    }


