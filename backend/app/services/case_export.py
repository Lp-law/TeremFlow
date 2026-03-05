"""
Smart Case Export: generate a single-case XLSX for operational use (read-only).
Uses overview-summary, fee events, retainer ledger, expenses, deductible summary, raw import.
"""

from __future__ import annotations

import datetime as dt
import io
import json
import re
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.models.case import Case
from app.services import cases as case_service
from app.services import expenses as expense_service
from app.services import fees as fee_service
from app.services import retainer as retainer_service
from app.services.unified import get_unified_summary


def _sanitize_filename_ref(ref: str) -> str:
    """Replace non-alphanumeric/underscore/hyphen with underscore; limit length."""
    s = re.sub(r"[^\w\-]", "_", str(ref).strip())
    return (s[:80] or "case") if s else "case"


def _fmt_dec(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, Decimal):
        return str(v)
    return str(v)


def _fmt_date(v: Any) -> str:
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10] if hasattr(v, "isoformat") else str(v)
    return str(v)


def build_case_export_xlsx(db: Session, case_id: int) -> tuple[bytes, str]:
    """
    Build XLSX workbook for one case. Returns (xlsx_bytes, suggested_filename).
    Read-only; no business logic or data changes.
    """
    case = case_service.get_case_if_not_deleted(db, case_id)
    if not case:
        return (b"", "")

    overview = case_service.build_case_overview_summary(db, case_id)
    unified = get_unified_summary(db, case)
    ledger = retainer_service.build_retainer_ledger(db, case_id=case_id)
    fee_events = fee_service.list_fee_events(db, case_id, include_deleted=False)
    fee_events_deleted = fee_service.list_fee_events(db, case_id, include_deleted=True)
    deleted_only = [e for e in fee_events_deleted if getattr(e, "deleted_at", None) is not None]
    expenses_list = expense_service.list_expenses(db, case_id)

    stages = case_service.get_latest_fee_stage_by_case_ids(db, [case.id])
    computed_stage = stages.get(case.id)
    procedure_stage_effective = case_service._effective_procedure_stage(case, computed_stage) or ""

    exported_at = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()
    # Remove default sheet after we create our first named sheet
    default_sheet = wb.active
    default_sheet.title = "Case"

    # --- Sheet: Case ---
    ws_case = wb["Case"]
    ws_case.append(["field", "value"])
    ws_case.append(["case_reference", case.case_reference or ""])
    ws_case.append(["case_name", case.case_name or ""])
    ws_case.append(["branch_name", case.branch_name or ""])
    ws_case.append(["status", case.status.value if case.status else ""])
    ws_case.append(["case_type", case.case_type.value if case.case_type else ""])
    ws_case.append(["open_date", _fmt_date(case.open_date)])
    ws_case.append(["procedure_stage_effective", procedure_stage_effective])
    ws_case.append(["exported_at", exported_at])

    # --- Sheet: Overview (unified model) ---
    ws_ov = wb.create_sheet("Overview")
    ws_ov.append(["field", "value"])
    if overview:
        ws_ov.append(["current_procedure_stage", overview.get("current_procedure_stage") or ""])
        fees = overview.get("fees") or {}
        ws_ov.append(["retainer_charged_to_date_ils", _fmt_dec(fees.get("retainer_charged_to_date_ils"))])
        ws_ov.append(["fees_by_stages_ils", _fmt_dec(fees.get("fees_by_stages_ils"))])
        ws_ov.append(["fee_diff_ils", _fmt_dec(fees.get("fee_diff_ils"))])
        ws_ov.append(["last_fee_event_date", fees.get("last_fee_event_date") or ""])
        ws_ov.append(["last_fee_event_amount", _fmt_dec(fees.get("last_fee_event_amount"))])
        ret = overview.get("retainer") or {}
        ws_ov.append(["charged_months_count", ret.get("charged_months_count") if ret.get("charged_months_count") is not None else ""])
        ws_ov.append(["monthly_gross_ils", _fmt_dec(ret.get("monthly_gross_ils"))])
        ws_ov.append(["retainer_is_frozen", ret.get("retainer_is_frozen") if ret.get("retainer_is_frozen") is not None else ""])
        ws_ov.append(["retainer_frozen_at", ret.get("retainer_frozen_at") or ""])
        exp = overview.get("expenses") or {}
        ws_ov.append(["total_expenses_ils", _fmt_dec(exp.get("total_expenses_ils"))])
        ded = overview.get("deductible") or {}
        ws_ov.append(["excess_total_ils", _fmt_dec(ded.get("excess_total_ils"))])
        ws_ov.append(["excess_remaining_ils", _fmt_dec(ded.get("excess_remaining_ils"))])

    # --- Sheet: Fees ---
    ws_fees = wb.create_sheet("Fees")
    headers_fees = [
        "event_date", "event_type", "amount_ils_gross", "retainer_credit_applied",
        "breakdown_new_codes", "breakdown_delta_total", "breakdown_adjustment", "breakdown_final_delta_total", "created_at",
    ]
    ws_fees.append(headers_fees)
    for e in fee_events:
        bj = e.breakdown_json or {}
        new_codes = bj.get("new_codes") or bj.get("codes")
        if isinstance(new_codes, list):
            new_codes_str = ",".join(str(c) for c in new_codes)
        else:
            new_codes_str = str(new_codes) if new_codes else ""
        ws_fees.append([
            _fmt_date(e.event_date),
            e.event_type.value if hasattr(e.event_type, "value") else str(e.event_type),
            _fmt_dec(e.computed_amount_ils_gross),
            _fmt_dec(e.amount_covered_by_credit_ils_gross),
            new_codes_str,
            _fmt_dec(bj.get("delta_total")),
            _fmt_dec(bj.get("adjustment")),
            _fmt_dec(bj.get("final_delta_total") or bj.get("final_total")),
            e.created_at.isoformat() if e.created_at else "",
        ])

    # --- Sheet: Deleted Fee Events (analytics) ---
    ws_del = wb.create_sheet("Deleted Fee Events")
    ws_del.append(["event_id", "event_date", "event_type", "amount_ils_gross", "delete_reason", "deleted_at"])
    for e in deleted_only:
        ws_del.append([
            str(e.id),
            _fmt_date(e.event_date),
            e.event_type.value if hasattr(e.event_type, "value") else str(e.event_type),
            _fmt_dec(e.computed_amount_ils_gross),
            getattr(e, "delete_reason", None) or "",
            e.deleted_at.isoformat() if getattr(e, "deleted_at", None) else "",
        ])

    # --- Sheet: Retainer ---
    ws_ret = wb.create_sheet("Retainer")
    ws_ret.append(["month", "row_type", "accrued_ils", "paid_ils", "running_credit_ils", "notes"])
    if ledger:
        cfg = ledger.get("config") or {}
        ws_ret.append(["", "config_monthly_base_net_ils", _fmt_dec(cfg.get("monthly_base_net_ils")), "", "", ""])
        ws_ret.append(["", "config_vat_pct", "", "", "", cfg.get("vat_pct") or ""])
        ws_ret.append(["", "config_monthly_gross_ils", _fmt_dec(cfg.get("monthly_gross_ils")), "", "", ""])
        ws_ret.append([ledger.get("anchor_date") or "", "anchor_date", "", "", "", ""])
        ws_ret.append([ledger.get("snapshot_through_month") or "", "snapshot_through_month", "", "", "", ""])
        ws_ret.append(["", "snapshot_paid_ils", _fmt_dec(ledger.get("snapshot_paid_ils")), "", "", ""])
        for row in ledger.get("rows") or []:
            ws_ret.append([
                row.get("month") or "",
                row.get("row_type") or "",
                _fmt_dec(row.get("accrued_ils")),
                _fmt_dec(row.get("paid_ils")),
                _fmt_dec(row.get("running_credit_ils")),
                row.get("notes") or "",
            ])

    # --- Sheet: Expenses ---
    ws_exp = wb.create_sheet("Expenses")
    ws_exp.append(["expense_date", "amount_ils_gross", "payer", "category", "supplier_name", "service_description", "demand_received_date", "attachment_url"])
    for ex in expenses_list:
        ws_exp.append([
            _fmt_date(ex.expense_date),
            _fmt_dec(ex.amount_ils_gross),
            ex.payer.value if hasattr(ex.payer, "value") else str(ex.payer),
            ex.category.value if hasattr(ex.category, "value") else str(ex.category),
            ex.supplier_name or "",
            (ex.service_description or "")[:500],
            _fmt_date(ex.demand_received_date),
            ex.attachment_url or "",
        ])

    # --- Sheet: Deductible (unified) ---
    ws_ded = wb.create_sheet("Deductible")
    ws_ded.append(["field", "value"])
    ws_ded.append(["excess_total_ils", _fmt_dec(unified.get("excess_total_ils"))])
    ws_ded.append(["retainer_charged_to_date_ils", _fmt_dec(unified.get("retainer_charged_to_date_ils"))])
    ws_ded.append(["expenses_total_ils", _fmt_dec(unified.get("expenses_total_ils"))])
    ws_ded.append(["fees_by_stages_ils", _fmt_dec(unified.get("fees_by_stages_ils"))])
    ws_ded.append(["excess_remaining_ils", _fmt_dec(unified.get("excess_remaining_ils"))])
    ws_ded.append(["fee_diff_ils", _fmt_dec(unified.get("fee_diff_ils"))])
    overrides = getattr(case, "manual_overrides_json", None) or {}
    if overrides:
        ws_ded.append(["", ""])
        ws_ded.append(["manual_overrides", json.dumps(overrides, ensure_ascii=False)])

    # --- Sheet: Raw Import ---
    ws_raw = wb.create_sheet("Raw Import")
    ws_raw.append(["key", "value"])
    raw = case.raw_import_fields_json or {}
    for k, val in sorted(raw.items()):
        if isinstance(val, (dict, list)):
            val_str = json.dumps(val, ensure_ascii=False)
        else:
            val_str = "" if val is None else str(val)
        ws_raw.append([str(k), val_str])
    if case.legacy_fee_text and str(case.legacy_fee_text).strip():
        ws_raw.append(["legacy_fee_text", str(case.legacy_fee_text).strip()])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    xlsx_bytes = buf.read()

    date_suffix = dt.date.today().strftime("%Y%m%d")
    ref_safe = _sanitize_filename_ref(case.case_reference or "case")
    filename = f"case_{ref_safe}_{date_suffix}.xlsx"

    return (xlsx_bytes, filename)
